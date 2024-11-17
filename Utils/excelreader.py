import openpyxl
import yaml


#gloabl variable 
columngetFrom_index=None
columnTarget_index=None

neededcell_rowindex=None
neededcell_colindex=None

def load_config(configpath):
    """ load the directory of file written in yaml"""
    with open (configpath,'r') as file :
        return yaml.safe_load(file)


    """
      get_column_index : get the column of a certain value  

    """   
def get_column_index(worksheet,value):
    col_index=None
    # Iterate over rows to find the matching value
    for row in worksheet.iter_rows(min_row=1,min_col=1, max_row=worksheet.max_row,values_only= True):
        #print(row)
        if value in row:
            col_index=row.index(value) # print the index in a tupel and is zero based 
            break
    if col_index !=None:
        return col_index+1

def get_row_index(worksheet,value):
    row_index=None
     # Iterate over rows to find the matching value
    for col in worksheet.iter_cols(values_only= True):
        if value in col:
            row_index=col.index(value) # print the index in a tupel and is zero based 
            break
    if row_index !=None:
        return row_index + 1
  

yamldata=load_config('Config\config.yaml')
workbook=openpyxl.load_workbook(yamldata['spreadsheet_name'],data_only=True)
worksheet= workbook[yamldata['sheet_name']]
# Get the value of a specific cell
cell_value = worksheet.cell(row=1, column=1).value
# print(cell_value)
col_index=None
# col_index= get_column_index(worksheet,0)
# print(col_index,end ="")

# row_index=get_row_index(worksheet,1,4)
# print(row_index)

neededcell_rowindex= get_row_index(worksheet,"yoyo")
neededcell_colindex =get_column_index(worksheet, yamldata['columnTarget'])

print(f'the row index is {neededcell_rowindex}')
print(neededcell_colindex)

if neededcell_colindex !=None and  neededcell_rowindex!=None:
    print(worksheet.cell(row=12,column=10).value)
else:
    print("There is no cell by this data")    